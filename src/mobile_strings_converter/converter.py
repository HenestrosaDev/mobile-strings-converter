import csv
import json
import os
import re
import warnings
from contextlib import redirect_stderr, redirect_stdout
from pathlib import Path
from typing import List, Tuple

import ezodf
import gspread
import openpyxl
import yaml
from arabic_reshaper import reshape
from bidi.algorithm import get_display
from console_style import ConsoleStyle
from fpdf import FPDF
from google.oauth2.credentials import Credentials
from lingua import LanguageDetectorBuilder
from PyPDF2 import PdfReader


def convert_strings(
    input_filepath: Path, output_filepath: Path, should_print_comments: bool
):
    """
    Extracts strings from the input file in either .xml or .strings format and converts
    them to the desired output file format. The output file format can be any of the
    following:

    Supported formats and corresponding extraction functions:
    - .csv: to_csv
    - .xlsx: to_sheet
    - .ods: to_sheet
    - .md: to_md
    - .json: to_json
    - .yaml: to_yaml
    - .html: to_html
    - .strings: to_ios
    - .xml: to_android
    - .pdf: to_pdf

    :param input_filepath: .strings or .xml file to extract the strings
    :type input_filepath: Path
    :param output_filepath: Name of the sheet to be generated
    :type output_filepath: Path
    :param should_print_comments: True if the user wants to print comments from
        .strings/.xml to the output file
    :type should_print_comments: bool
    """

    strings = get_strings(input_filepath, should_print_comments)

    if output_filepath:
        conversion_functions = {
            ".csv": to_csv,
            ".xlsx": to_sheet,
            ".ods": to_sheet,
            ".md": to_md,
            ".json": to_json,
            ".yaml": to_yaml,
            ".html": to_html,
            ".strings": to_ios,
            ".xml": to_android,
            ".pdf": to_pdf,
        }

        if output_filepath.suffix in conversion_functions:
            conversion_functions[output_filepath.suffix](strings, output_filepath)

            print(
                f"{ConsoleStyle.GREEN}Data successfully written to {output_filepath}"
                f"{ConsoleStyle.END}"
            )
        else:
            raise ValueError(
                f"{ConsoleStyle.YELLOW}File type not supported. Feel free to create "
                f"an issue here (https://github.com/HenestrosaConH/mobile-strings"
                f"-converter/issues) if you want the file type to be supported by the "
                f"package.{ConsoleStyle.END}"
            )


def get_strings(
    input_filepath: Path, should_print_comments: bool
) -> List[Tuple[str, str]]:
    """
    Extracts strings from various file formats based on the file extension.

    Supported formats and corresponding extraction functions:
    - .csv: get_strings_from_csv
    - .xlsx: get_strings_from_xlsx
    - .ods: get_strings_from_ods
    - .md: get_strings_from_md
    - .json: get_strings_from_json
    - .yaml: get_strings_from_yaml
    - .html: get_strings_from_html
    - .strings: get_strings_from_ios
    - .xml: get_strings_from_xml
    - .pdf: get_strings_from_pdf

    If the input file format is .strings or .xml, additional options are available:
    - should_print_comments: If True, includes comments in the extracted strings.

    :param input_filepath: Path to the input file.
    :type input_filepath: Path
    :param should_print_comments: True if comments should be included (for .strings and
        .xml files), False otherwise.
    :type should_print_comments: bool
    :return: A list of tuples containing extracted strings and their corresponding values.
    :rtype: List[Tuple[str, str]]
    """

    conversion_functions = {
        ".csv": get_strings_from_csv,
        ".xlsx": get_strings_from_xlsx,
        ".ods": get_strings_from_ods,
        ".md": get_strings_from_md,
        ".json": get_strings_from_json,
        ".yaml": get_strings_from_yaml,
        ".html": get_strings_from_html,
        ".strings": get_strings_from_ios,
        ".xml": get_strings_from_xml,
        ".pdf": get_strings_from_pdf,
    }

    if input_filepath.suffix in [".strings", ".xml"]:
        return conversion_functions[input_filepath.suffix](
            input_filepath, should_print_comments
        )
    else:
        return conversion_functions[input_filepath.suffix](input_filepath)


def to_google_sheets(
    input_filepath: Path,
    sheet_name: str,
    credentials_filepath: Path,
    should_print_comments: bool,
):
    """
    Creates a Google spreadsheet with the extracted strings from the input filepath

    :param input_filepath: .strings or .xml file to extract the strings
    :type input_filepath: Path
    :param sheet_name: Name of the sheet to be generated
    :type sheet_name: str
    :param credentials_filepath: Path to the service_account.json in order to be able
        to create the sheet in the user's Google account
    :type credentials_filepath: Path
    :param should_print_comments: True if the user wants to print comments from
        .strings/.xml to the sheet
    :type should_print_comments: bool
    """

    strings = get_strings(input_filepath, should_print_comments)

    # Authenticate with Google Sheets API
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    credentials = Credentials.from_service_account_file(credentials_filepath, scope)
    client = gspread.authorize(credentials)

    # Open a new sheet or an existing one
    sheet = client.open(sheet_name).sheet1

    # Clear the existing data in the sheet
    sheet.clear()

    # Write the data to the sheet
    for string in strings:
        sheet.append_row(string)


def to_csv(strings: List[str], output_filepath: Path):
    """
    Formats strings to a .csv file

    :param strings: Strings extracted from a .strings or .xml file
    :type strings: List[str]
    :param output_filepath: The path where the generated file will be saved.
    :type output_filepath: Path
    """

    with open(output_filepath, "w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)

        # Write the header row
        header = ["name", "value"]
        writer.writerow(header)

        # Write the data to the file
        for name, value in strings:
            writer.writerow([name, value])


def to_sheet(strings: List[str], output_filepath: Path):
    """
    Formats strings to a .xlsx / .ods file

    :param strings: Strings extracted from a .strings or .xml file
    :type strings: List[str]
    :param output_filepath: The path where the generated file will be saved.
    :type output_filepath: Path
    """

    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Create a new sheet
    sheet = workbook.active

    # Write the header row
    sheet.cell(row=1, column=1, value="NAME")
    sheet.cell(row=1, column=2, value="VALUE")

    # Write the data to the sheet
    for i, (name, value) in enumerate(strings, start=2):
        sheet.cell(row=i, column=1, value=name)
        sheet.cell(row=i, column=2, value=value)

    # Save the file
    workbook.save(output_filepath)


def to_json(strings: List[str], output_filepath: Path):
    """
    Formats strings to a .json file

    :param strings: Strings extracted from a .strings or .xml file
    :type strings: List[str]
    :param output_filepath: The path where the generated file will be saved.
    :type output_filepath: Path
    """

    # Create a list of dictionaries to store the data
    data_list = []
    for name, value in strings:
        data_list.append({"name": name, "value": value})

    # Write the data to the JSON file
    with open(output_filepath, "w", encoding="utf-8") as file:
        json.dump(data_list, file, ensure_ascii=False, indent=2)


def to_yaml(strings: List[str], output_filepath: Path):
    """
    Formats strings to a .yaml file

    :param strings: Strings extracted from a .strings or .xml file
    :type strings: List[str]
    :param output_filepath: The path where the generated file will be saved.
    :type output_filepath: Path
    """

    # Convert the data to a dictionary
    strings_dict = {name: value for name, value in strings}

    # Write the data to the YAML file
    with open(output_filepath, "w", encoding="utf-8") as file:
        yaml.dump(strings_dict, file, default_flow_style=False, allow_unicode=True)


def to_html(strings: List[str], output_filepath: Path):
    """
    Formats strings to a .html file

    :param strings: Strings extracted from a .strings or .xml file
    :type strings: List[str]
    :param output_filepath: The path where the generated file will be saved.
    :type output_filepath: Path
    """

    # Create an HTML file
    with open(output_filepath, "w", encoding="utf-8") as file:
        file.write("<head>\n")
        file.write('\t<meta charset="UTF-8">\n')
        file.write("</head>\n")
        file.write("<table>\n")
        file.write("\t<thead>\n")
        file.write("\t\t<tr>\n")
        file.write("\t\t\t<th>NAME</th>\n")
        file.write("\t\t\t<th>VALUE</th>\n")
        file.write("\t\t</tr>\n")
        file.write("\t</thead>\n")
        file.write("\t<tbody>\n")

        # Write the data to the HTML file
        for name, value in strings:
            file.write("\t\t<tr>\n")
            file.write(f"\t\t\t<td>{name}</td>\n")
            file.write(f"\t\t\t<td>{value}</td>\n")
            file.write("\t\t</tr>\n")

        file.write("\t</tbody>\n")
        file.write("</table>\n")


def to_ios(strings: List[str], output_filepath: Path):
    """
    Formats strings to a .strings file

    :param strings: Strings extracted from a .strings or .xml file
    :type strings: List[str]
    :param output_filepath: The path where the generated file will be saved.
    :type output_filepath: Path
    """

    with open(output_filepath, "w", encoding="utf-8") as file:
        for string in strings:
            file.write(f'"{string[0]}" = "{string[1]}";\n')


def to_android(strings: List[str], output_filepath: Path):
    """
    Formats strings to a .xml file

    :param strings: Strings extracted from a .strings or .xml file
    :type strings: List[str]
    :param output_filepath: The path where the generated file will be saved.
    :type output_filepath: Path
    """

    with open(output_filepath, "w", encoding="utf-8") as file:
        file.write("<resources>\n")
        for string in strings:
            file.write(f'\t<string name="{string[0]}">{string[1]}</string>\n')

        file.write("</resources>")


def to_pdf(strings: List[str], output_filepath: Path):
    """
    Formats strings to a .pdf file

    :param strings: Strings extracted from a .strings or .xml file
    :type strings: List[str]
    :param output_filepath: The path where the generated file will be saved.
    :type output_filepath: Path
    """

    # Ignore the following warning when adding a font already added:
    # UserWarning: Core font or font already added 'dejavusanscondensed': doing nothing
    warnings.filterwarnings("ignore", category=UserWarning)

    def add_font(font_name, size=12):
        root_dir = Path(__file__).parent
        pdf.add_font(fname=str(root_dir / f"assets/fonts/{font_name}.ttf"))
        pdf.set_font(font_name, size=size)

    # Create a new PDF file
    pdf = FPDF(orientation="P", format="A4")
    pdf.add_page()
    pdf.set_font("Arial", "B", 12)

    # Cell properties
    c_width = 95
    c_height = 10

    # Add headers to table
    pdf.cell(c_width, c_height, "NAME", border=1)
    pdf.cell(c_width, c_height, "VALUE", border=1)
    pdf.ln()

    detector = (
        LanguageDetectorBuilder.from_all_languages()
        .with_preloaded_language_models()
        .build()
    )

    # Add table data
    # https://stackoverflow.com/questions/53526311/fpdf-multicell-same-height
    for i, string in enumerate(strings):
        x = pdf.get_x()
        y = pdf.get_y()

        max_height = 0
        cells_in_row = 2

        for j in range(cells_in_row):
            language_code = None
            try:
                if j % 2 == 0:  # Prevents 'name' language detection
                    add_font("DejaVuSansCondensed")
                else:
                    language_code = detector.detect_language_of(
                        string[j]
                    ).iso_code_639_1.name.lower()

                    if language_code in [
                        "bn",  # Bengali
                        "hi",  # Hindi
                        "kn",  # Kannada
                        "ml",  # Malayalam
                        "mr",  # Marathi
                        "or",  # Oriya
                        "bo",  # Tibetan
                    ]:
                        add_font("gargi")
                    elif language_code == "gu":  # Gujarati
                        add_font("Aakar")
                    elif language_code == "te":  # Telugu
                        add_font("AnekTelugu-VariableFont_wdth,wght")
                    elif language_code == "ta":  # Tamil
                        add_font("latha")
                    elif language_code == "pa":  # Punjabi, Panjabi
                        add_font("Gurvetica_a8_Heavy")
                    elif language_code == "zh" or language_code == "ja":
                        # Chinese or Japanese
                        add_font("fireflysung")
                    elif language_code == "ko":  # Korean
                        add_font("Eunjin")
                    elif language_code == "th":  # Thai
                        add_font("Waree")
                    else:
                        add_font("DejaVuSansCondensed")

                if language_code in [
                    # RTL languages
                    "ar",  # Arabic
                    "he",  # Hebrew
                    "dv",  # Dhivehi
                    "ku",  # Kurdish (sorani)
                    "ps",  # Pashto
                    "fa",  # Persian
                    "sd",  # Sindhi
                    "ur",  # Urdu
                    "ug",  # Uyghur
                    "yi",  # Yiddish
                ]:
                    pdf.multi_cell(c_width, c_height, get_display(reshape(string[j])))
                else:
                    pdf.multi_cell(c_width, c_height, string[j])

                if pdf.get_y() - y > max_height:
                    max_height = pdf.get_y() - y

                pdf.set_xy(x + (c_width * (j + 1)), y)
            except (Exception,):
                with open(
                    output_filepath.parent / f"{output_filepath.stem}-errors.txt",
                    "a",
                    encoding="utf-8",
                ) as f:
                    f.write(f"{string[1]} not supported\n")

        for j in range(cells_in_row + 1):
            pdf.line(x + c_width * j, y, x + c_width * j, y + max_height)

        pdf.line(x, y, x + c_width * cells_in_row, y)
        pdf.line(x, y + max_height, x + c_width * cells_in_row, y + max_height)

        pdf.ln()

        if (
            i < len(strings) - 1
            and pdf.get_y() + (max_height * cells_in_row) > pdf.h - 10
        ):
            pdf.add_page()

    # Inside this context manager, all output to stdout and stderr will be suppressed
    # This is done because in Windows, the following exception is raised if the
    # strings.xml file contains unsupported characters:
    #
    # UnicodeEncodeError: 'charmap'
    # codec can't encode characters in position 0-9: character maps to <undefined>
    with open(os.devnull, "w") as devnull:
        with redirect_stdout(devnull), redirect_stderr(devnull):
            # Save the PDF file
            pdf.output(str(output_filepath))


def to_md(strings: List[str], output_filepath: Path):
    """
    Formats strings to a .md file

    :param strings: Strings extracted from a .strings or .xml file
    :type strings: List[str]
    :param output_filepath: The path where the generated file will be saved.
    :type output_filepath: Path
    """

    with open(output_filepath, "w", encoding="utf-8") as f:
        # Write each string to the Markdown file in a table format
        f.write("| NAME | VALUE |\n")
        f.write("| ----------- | ----------- |\n")
        for name, translation in strings:
            f.write(f"| {name} | {translation} |\n")


# GET STRINGS FROM


def get_strings_from_csv(csv_filepath: Path):
    """
    Extract data from a CSV file with NAME and VALUE columns and return it as a
    list of tuples.

    :param csv_filepath: The path to the input CSV file.
    :type csv_filepath: Path
    :return: A list of tuples where each tuple contains a NAME and VALUE.
    :rtype: List[Tuple[str, str]]
    """

    # Initialize a list to hold the tuples
    data = []

    # Open the CSV file and read its contents
    with open(csv_filepath, "r", newline="", encoding="utf-8") as file:
        csv_reader = csv.reader(file)
        next(csv_reader)  # Skip the header row

        # Iterate over the rows in the CSV file
        for row in csv_reader:
            name, value = row
            data.append((name, value))

    return data


def get_strings_from_xlsx(sheet_filepath: Path):
    """
    Extract data from an Excel file with NAME and VALUE columns and return it as a list
    of tuples.

    :param sheet_filepath: The path to the input Excel file.
    :type sheet_filepath: str
    :return: A list of tuples where each tuple contains a NAME and VALUE.
    :rtype: List[Tuple[str, str]]
    """

    # Load the workbook and select the active sheet
    workbook = openpyxl.load_workbook(sheet_filepath)
    sheet = workbook.active

    # Initialize a list to hold the tuples
    data = []

    # Iterate over the rows in the sheet starting from the second row to skip the header
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, value = row
        data.append((name, value))

    return data


def get_strings_from_ods(ods_filepath: Path) -> List[Tuple[str, str]]:
    """
    Extract data from an OpenDocument Spreadsheet (ODS) file with NAME and VALUE columns
    and return it as a list of tuples.

    :param ods_filepath: The path to the input ODS file.
    :type ods_filepath: str
    :return: A list of tuples where each tuple contains a NAME and VALUE.
    :rtype: List[Tuple[str, str]]
    """

    # Initialize a list to hold the tuples
    data = []

    # Load the ODS file
    doc = ezodf.opendoc(ods_filepath)

    # Get the first sheet
    sheet = doc.sheets[0]

    # Iterate over the rows in the sheet
    for row in sheet.rows():
        # Extract NAME and VALUE from each row
        name, value = [cell.value for cell in row[:2]]
        data.append((name, value))

    return data


def get_strings_from_md(
    md_filepath: Path, delimiter: str = "|"
) -> List[Tuple[str, str]]:
    """
    Extract data from a Markdown file with a table containing NAME and VALUE columns and
    return it as a list of tuples.

    :param md_filepath: The path to the input Markdown file.
    :type md_filepath: Path
    :param delimiter: The delimiter used in the Markdown table, defaults to '|'.
    :type delimiter: str, optional
    :return: A list of tuples where each tuple contains a NAME and VALUE.
    :rtype: List[Tuple[str, str]]
    """

    # Initialize a list to hold the tuples
    data = []

    # Open the Markdown file and read its contents
    with open(md_filepath, "r", encoding="utf-8") as file:
        lines = file.readlines()

    # Find the start and end indices of the table
    start_index = None
    end_index = None
    for i, line in enumerate(lines):
        if line.strip().startswith("|"):
            start_index = i
            break
    for i in range(len(lines) - 1, -1, -1):
        if lines[i].strip().startswith("|"):
            end_index = i
            break

    # Extract data from the table, skipping the first two lines (header)
    if start_index is not None and end_index is not None:
        for row in lines[start_index + 2 : end_index + 1]:
            # Split the line by the delimiter and extract NAME and VALUE
            parts = row.strip().strip(delimiter).split(delimiter)
            if len(parts) >= 2:
                name, value = parts[:2]
                data.append((name.strip(), value.strip()))

    return data


def get_strings_from_json(json_filepath: Path) -> List[Tuple[str, str]]:
    """
    Extract data from a JSON file with objects containing NAME and VALUE fields and
    return it as a list of tuples.

    :param json_filepath: The path to the input JSON file.
    :type json_filepath: Path
    :return: A list of tuples where each tuple contains a NAME and VALUE.
    :rtype: List[Tuple[str, str]]
    """

    # Initialize a list to hold the tuples
    data = []

    # Open the JSON file and load its contents
    with open(json_filepath, "r", encoding="utf-8") as file:
        json_data = json.load(file)

    # Iterate over each object in the JSON data
    for record in json_data:
        if "name" in record and "value" in record:
            data.append((record["name"], record["value"]))

    return data


def get_strings_from_yaml(yaml_filepath: Path) -> List[Tuple[str, str]]:
    """
    Extract data from a YAML file with objects containing NAME and VALUE fields and
    return it as a list of tuples.

    :param yaml_filepath: The path to the input YAML file.
    :type yaml_filepath: Path
    :return: A list of tuples where each tuple contains a NAME and VALUE.
    :rtype: List[Tuple[str, str]]
    """

    # Initialize a list to hold the tuples
    data = []

    # Open the YAML file and load its contents
    with open(yaml_filepath, "r", encoding="utf-8") as file:
        yaml_data = yaml.safe_load(file)

    # Iterate over each key-value pair in the YAML data
    for key, value in yaml_data.items():
        data.append((key, value))

    return data


def get_strings_from_html(html_filepath: Path) -> List[Tuple[str, str]]:
    """
    Extract data from an HTML file with a table containing NAME and VALUE columns and
    return it as a list of tuples.

    :param html_filepath: The path to the input HTML file.
    :type html_filepath: Path
    :return: A list of tuples where each tuple contains a NAME and VALUE.
    :rtype: List[Tuple[str, str]]
    """

    # Initialize a list to hold the tuples
    data = []

    # Open the HTML file and read its contents
    with open(html_filepath, "r", encoding="utf-8") as file:
        html_content = file.read()

    # Find the start and end indices of the table
    table_start = html_content.find("<table")
    table_end = html_content.find("</table>", table_start)

    # Extract data from the table if it exists
    if table_start != -1 and table_end != -1:
        table_content = html_content[table_start:table_end]
        rows = re.findall(r"<tr[^>]*>(.*?)</tr>", table_content, re.DOTALL)

        # Extract data from each row
        for row in rows:
            cells = re.findall(r"<td[^>]*>(.*?)</td>", row, re.DOTALL)
            if len(cells) >= 2:
                name = re.sub(r"<.*?>", "", cells[0].strip())
                value = re.sub(r"<.*?>", "", cells[1].strip())
                data.append((name, value))

    return data


def get_strings_from_ios(
    ios_filepath: Path, should_print_comments: bool
) -> List[Tuple[str, str]]:
    """
    Get strings from the .strings or .xml file.

    :param ios_filepath: .strings or .xml file to extract the strings
    :type ios_filepath: Path
    :param should_print_comments: True if the user wants to print comments from
        the .strings to the output file
    :type should_print_comments: bool
    :return: A list of tuples where each tuple contains a NAME and VALUE.
    :rtype: List[Tuple[str, str]]
    """

    if should_print_comments:
        pattern = r'"(.*?)"\s*=\s*"((?:[^"\\]|\\.)*)"\s*;'
    else:
        pattern = r'^(?!\s*//)\s*"(.+?)"\s*=\s*"((?:[^"\\]|\\.)*)"\s*;'

    # Open the strings file
    with open(ios_filepath, "r", encoding="utf-8") as file:
        strings_data = file.read()

    # Extract the strings using a regular expression
    strings = re.findall(pattern, strings_data, re.MULTILINE)

    if len(strings) >= 1:
        return strings
    else:
        raise ValueError("The file provided is not a valid .strings file.")


def get_strings_from_xml(
    xml_filepath: Path, should_print_comments: bool
) -> List[Tuple[str, str]]:
    """
    Get strings from the .strings or .xml file.

    :param xml_filepath: .strings or .xml file to extract the strings
    :type xml_filepath: Path
    :param should_print_comments: True if the user wants to print comments from
        the .strings to the output file
    :type should_print_comments: bool
    :return: A list of tuples where each tuple contains a NAME and VALUE.
    :rtype: List[Tuple[str, str]]
    """

    if should_print_comments:
        pattern = r'<string name="(.*?)">(.*?)</string>'
    else:
        pattern = r'^(?!\s*<!--)\s*<string name="(.*?)">(.*?)</string>(?!\s*-->)'

    # Open the strings file
    with open(xml_filepath, "r", encoding="utf-8") as file:
        strings_data = file.read()

    # Extract the strings using a regular expression
    strings = re.findall(pattern, strings_data, re.MULTILINE)

    if len(strings) >= 1:
        return strings
    else:
        raise ValueError("The file provided is not a valid .xml file.")


def get_strings_from_pdf(pdf_filepath: Path) -> List[Tuple[str, str]]:
    """
    Extract data from a PDF file with a table containing NAME and VALUE columns and
    return it as a list of tuples.

    :param pdf_filepath: The path to the input PDF file.
    :type pdf_filepath: Path
    :return: A list of tuples where each tuple contains a NAME and VALUE.
    :rtype: List[Tuple[str, str]]
    """
    # Initialize a list to hold the tuples
    data = []

    # Create a PdfReader object
    pdf_reader = PdfReader(pdf_filepath)

    # Extract text from each page
    for page in pdf_reader.pages:
        text = page.extract_text()

        # Find patterns for table rows
        rows = text.split("\n")

        # Skip the header
        rows = rows[1:]

        for row in rows:
            match = re.match(r"(\w+)\s+(.*)", row.strip())
            if match:
                name, value = match.groups()
                data.append((name.strip(), value.strip()))

    return data
