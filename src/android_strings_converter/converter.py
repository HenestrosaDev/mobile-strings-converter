import csv
import json
import re
from pathlib import Path

import gspread
import openpyxl
import yaml
from arabic_reshaper import reshape
from bidi.algorithm import get_display
from fpdf import FPDF
from google.oauth2.credentials import Credentials
from lingua import LanguageDetectorBuilder


def get_xml_strings(xml_filepath: Path):
    # Open the string.xml file
    with open(xml_filepath, "r", encoding="utf-8") as file:
        xml_data = file.read()

    # Extract the "name" attribute and content of the tag
    pattern = r'<string name="(.*?)">(.*?)</string>'
    strings = re.findall(pattern, xml_data)

    if len(strings) >= 1:
        return strings
    else:
        raise ValueError("The XML file provided is not a valid Android strings file.")


def to_google_sheets(xml_filepath: Path, sheet_name: str, credentials_filepath: Path):
    strings = get_xml_strings(xml_filepath)

    # Authenticate with Google Sheets API
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    credentials = Credentials.from_service_account_file(credentials_filepath, scope)
    client = gspread.authorize(credentials)

    # Open a new sheet or an existing one
    sheet = client.open(sheet_name).sheet1

    # Clear the existing data in the sheet
    sheet.clear()

    # Write the data to the sheet
    for string in strings:
        sheet.append_row(string)


def to_csv(xml_filepath: Path, csv_filepath: Path):
    strings = get_xml_strings(xml_filepath)

    # Create a CSV file
    with open(csv_filepath, "w", encoding="utf-8", newline="") as file:
        writer = csv.writer(file)

        # Write the header row
        header = ["name", "value"]
        writer.writerow(header)

        # Write the data to the file
        for name, value in strings:
            writer.writerow([name, value])


def to_xlsx_ods(xml_filepath: Path, xlsx_filepath: Path):
    strings = get_xml_strings(xml_filepath)

    # Create a new workbook
    workbook = openpyxl.Workbook()

    # Create a new sheet
    sheet = workbook.active

    # Write the header row
    sheet.cell(row=1, column=1, value="NAME")
    sheet.cell(row=1, column=2, value="VALUE")

    # Write the data to the sheet
    for i, (name, value) in enumerate(strings, start=2):
        sheet.cell(row=i, column=1, value=name)
        sheet.cell(row=i, column=2, value=value)

    # Save the file
    workbook.save(xlsx_filepath)


def to_json(xml_filepath: Path, json_filepath: Path):
    strings = get_xml_strings(xml_filepath)

    # Create a list of dictionaries to store the data
    data_list = []
    for name, value in strings:
        data_list.append({"name": name, "value": value})

    # Write the data to the JSON file
    with open(json_filepath, "w", encoding="utf-8") as file:
        json.dump(data_list, file, ensure_ascii=False, indent=2)


def to_yaml(xml_filepath: Path, yaml_filepath: Path):
    strings = get_xml_strings(xml_filepath)

    # Convert the data to a dictionary
    strings_dict = {name: value for name, value in strings}

    # Write the data to the YAML file
    with open(yaml_filepath, "w", encoding="utf-8") as file:
        yaml.dump(strings_dict, file, default_flow_style=False, allow_unicode=True)


def to_html(xml_filepath: Path, html_filepath: Path):
    strings = get_xml_strings(xml_filepath)

    # Create an HTML file
    with open(html_filepath, "w", encoding="utf-8") as file:
        file.write("<head>\n")
        file.write('\t<meta charset="UTF-8">\n')
        file.write("</head>\n")
        file.write("<table>\n")
        file.write("\t<thead>\n")
        file.write("\t\t<tr>\n")
        file.write("\t\t\t<th>NAME</th>\n")
        file.write("\t\t\t<th>VALUE</th>\n")
        file.write("\t\t</tr>\n")
        file.write("\t</thead>\n")
        file.write("\t<tbody>\n")

        # Write the data to the HTML file
        for name, value in strings:
            file.write("\t\t<tr>\n")
            file.write(f"\t\t\t<td>{name}</td>\n")
            file.write(f"\t\t\t<td>{value}</td>\n")
            file.write("\t\t</tr>\n")

        file.write("\t</tbody>\n")
        file.write("</table>\n")


def to_ios(xml_filepath: Path, localizable_filepath: Path):
    strings = get_xml_strings(xml_filepath)

    with open(localizable_filepath, "w", encoding="utf-8") as file:
        for string in strings:
            file.write(f'"{string[0]}" = "{string[1]}";\n')


def to_pdf(xml_filepath: Path, pdf_filepath: Path):
    def add_font(font_name, size=12):
        root_dir = Path(__file__).parent
        pdf.add_font(fname=str(root_dir / f"assets/fonts/{font_name}.ttf"))
        pdf.set_font(font_name, size=size)

    strings = get_xml_strings(xml_filepath)

    # Create a new PDF file
    pdf = FPDF(orientation="P", format="A4")
    pdf.add_page()
    pdf.set_font("Arial", "B", 12)

    # Cell properties
    c_width = 95
    c_height = 10

    # Add headers to table
    pdf.cell(c_width, c_height, "NAME", border=1)
    pdf.cell(c_width, c_height, "VALUE", border=1)
    pdf.ln()

    detector = (
        LanguageDetectorBuilder.from_all_languages()
        .with_preloaded_language_models()
        .build()
    )

    # Add table data
    # https://stackoverflow.com/questions/53526311/fpdf-multicell-same-height
    for i, string in enumerate(strings):
        x = pdf.get_x()
        y = pdf.get_y()

        max_height = 0
        cells_in_row = 2

        for j in range(cells_in_row):
            language_code = None
            try:
                if j % 2 == 0:  # Prevents 'name' language detection
                    add_font("DejaVuSansCondensed")
                else:
                    language_code = detector.detect_language_of(
                        string[j]
                    ).iso_code_639_1.name.lower()

                    if language_code in [
                        "bn",  # Bengali
                        "hi",  # Hindi
                        "kn",  # Kannada
                        "ml",  # Malayalam
                        "mr",  # Marathi
                        "or",  # Oriya
                        "bo",  # Tibetan
                    ]:
                        add_font("gargi")
                    elif language_code == "gu":  # Gujarati
                        add_font("Aakar")
                    elif language_code == "te":  # Telugu
                        add_font("AnekTelugu-VariableFont_wdth,wght")
                    elif language_code == "ta":  # Tamil
                        add_font("latha")
                    elif language_code == "pa":  # Punjabi, Panjabi
                        add_font("Gurvetica_a8_Heavy")
                    elif language_code == "zh" or language_code == "ja":
                        # Chinese or Japanese
                        add_font("fireflysung")
                    elif language_code == "ko":  # Korean
                        add_font("Eunjin")
                    elif language_code == "th":  # Thai
                        add_font("Waree")
                    else:
                        add_font("DejaVuSansCondensed")

                if language_code in [
                    # RTL languages
                    "ar",  # Arabic
                    "he",  # Hebrew
                    "dv",  # Dhivehi
                    "ku",  # Kurdish (sorani)
                    "ps",  # Pashto
                    "fa",  # Persian
                    "sd",  # Sindhi
                    "ur",  # Urdu
                    "ug",  # Uyghur
                    "yi",  # Yiddish
                ]:
                    pdf.multi_cell(c_width, c_height, get_display(reshape(string[j])))
                else:
                    pdf.multi_cell(c_width, c_height, string[j])

                if pdf.get_y() - y > max_height:
                    max_height = pdf.get_y() - y

                pdf.set_xy(x + (c_width * (j + 1)), y)
            except (Exception,):
                print(f"{string[1].encode('utf-8')} not supported")
                pass

        for j in range(cells_in_row + 1):
            pdf.line(x + c_width * j, y, x + c_width * j, y + max_height)

        pdf.line(x, y, x + c_width * cells_in_row, y)
        pdf.line(x, y + max_height, x + c_width * cells_in_row, y + max_height)

        pdf.ln()

        if (
            i < len(strings) - 1
            and pdf.get_y() + (max_height * cells_in_row) > pdf.h - 10
        ):
            pdf.add_page()

    # Save the PDF file
    pdf.output(str(pdf_filepath))


def to_md(xml_filepath: Path, md_filepath: Path):
    strings = get_xml_strings(xml_filepath)

    # Create a new Markdown file
    with open(md_filepath, "w", encoding="utf-8") as f:
        # Write each string to the Markdown file in a table format
        f.write("| NAME | VALUE |\n")
        f.write("| ----------- | ----------- |\n")
        for name, translation in strings:
            f.write(f"| {name} | {translation} |\n")
