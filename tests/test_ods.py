import unittest
from pathlib import Path

import openpyxl
import pandas as pd
from base_tests import BaseTests
from mobile_strings_converter.converter import convert_strings


class TestToOds(BaseTests.ConvertToTest):
    def setUp(self):
        super().setUp()
        self.file_name = "strings.ods"

    # Overrides
    def _converter_writes_correct_data(
        self,
        template_filepath: Path,
        input_filepath: Path,
        with_comments: bool,
    ):
        convert_strings(input_filepath, self.output_filepath, with_comments)

        # Load the two Excel files into pandas dataframes
        df1 = pd.read_excel(template_filepath)
        df2 = pd.read_excel(self.output_filepath)

        # Compare the dataframes to check if they are equal
        self.assertTrue(df1.equals(df2))


class TestFromOds(BaseTests.ConvertFromTest):
    def setUp(self):
        super().setUp()
        self.file_name = "strings.xlsx"

    # Overriding method
    def _converter_writes_correct_data(
        self, template_filepath: Path, input_filepath: Path
    ):
        convert_strings(input_filepath, self.output_filepath)

        wb1 = openpyxl.load_workbook(self.output_filepath)
        wb2 = openpyxl.load_workbook(template_filepath)

        # Compare sheet names
        sheets1 = wb1.sheetnames
        sheets2 = wb2.sheetnames

        self.assertEqual(sheets1, sheets2, "Sheet names differ")

        for sheet_name in sheets1:
            ws1 = wb1[sheet_name]
            ws2 = wb2[sheet_name]

            # Compare dimensions
            self.assertEqual(
                ws1.max_row, ws2.max_row, f"Sheet '{sheet_name}' row count differs"
            )
            self.assertEqual(
                ws1.max_column,
                ws2.max_column,
                f"Sheet '{sheet_name}' column count differs",
            )

            # Compare cell values
            for row in ws1.iter_rows():
                for cell in row:
                    value1 = cell.value
                    value2 = ws2[cell.coordinate].value

                    self.assertEqual(
                        value1,
                        value2,
                        f"Sheet '{sheet_name}' cell {cell.coordinate} differs: {value1} != {value2}",
                    )


if __name__ == "__main__":
    unittest.main()
